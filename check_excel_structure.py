#!/usr/bin/env python3

import openpyxl
import sys

excel_path = r"c:\laragon\www\survey\storage\app\psgc\PSGC-3Q-2025-Publication-Datafile.xlsx"

print("Reading Excel file structure...")

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

print(f"\nSheet name: {ws.title}")
print(f"Max row: {ws.max_row}")
print(f"Max column: {ws.max_column}")

print("\nFirst 15 rows (with column headers):")
for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
    print(f"Row {idx}: {row}")
